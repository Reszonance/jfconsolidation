from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.colors import Color

# keep this independent of shipping_objects.py

import os

  # returns workbook 
def initialize_workbook(filepath, overwrite):
  if not os.path.exists(filepath):
    wb = Workbook()  # name is given when saving
  else:
    if overwrite:
        os.remove(filepath)
        # appearing right in local files
        wb = Workbook()
    else:
      print(f'loading previous workbook, no overwrite. workbook filepath: {filepath}')
      wb = load_workbook(filepath)
  
  return wb

def save_data(wb, name, file_path, overwrite):
    # --------------------- SAVE DATA
    # check if workbook already exists
  if not overwrite:
    name = create_new_copy(name, 0)
      
  wb.save(file_path)
  return wb.path

def create_new_copy(sheet_name, copy_version):
    if (os.path.exists(sheet_name)):
        # find index of extension
      extension = sheet_name.split('.')[-1]
      if extension == 'xlsx':
          ind = sheet_name.rfind('.')
          """
          # this stuff accounts for multiple copies
          sn = sheet_name.replace(".xlsx", "")
          sn = sn.replace(sheet_name.replace(".xlsx", ""), "")
          if sn != "":
              # already exists a copy
            if create_new_copy(f"{sheet_name}({copy_version + 1})") != -1:
               pass
            
            """
          sheet_name = sheet_name[:ind]+'(1)'+sheet_name[ind:]
          return sheet_name
    return -1

def populate_grades(wb):
  ws = wb.active  # gets default sheet
  ws.title = "Grades"
    # do stuff here

    # region data
  data = {
    "Joe": {
      "math": 65,
      "science": 78,
      "english": 98,
      "gym": 89
    },
    "Bill": {
      "math": 55,
      "science": 72,
      "english": 87,
      "gym": 95
    },
    "Tim": {
      "math": 100,
      "science": 45,
      "english": 75,
      "gym": 92
    },
    "Sally": {
      "math": 30,
      "science": 25,
      "english": 45,
      "gym": 100
    },
    "Jane": {
      "math": 100,
      "science": 100,
      "english": 100,
      "gym": 60
    }
  }

    # endregion

  headings = ['Name'] + list(data['Joe'].keys())
  ws.append(headings)

  for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)


  for col in range(2, len(data['Joe']) + 2):
    srow = len(data)+2
    subject_num = len(data[person].values())
    char = get_column_letter(col)
    ws.cell(srow, col).value = f"=SUM({char+'2'}:{char+str(2+subject_num)})/{len(data)}"

    # bold headings
  for col in range(1, 6):
    ws.cell(1, col).font = Font(bold = True, color="0099CCFF")

def populate_shipment(wb, data):
  """
  Writes data to excel sheet given as wb.

  Parameters:
  ----------
  wb: Workbook
  data: dict
    Contains data on shipment.

  Returns
  ----------
  None
  """
  ws = wb.active  # gets default sheet
  ws.title = "Customer Info"
  
  print(f'\nPOPULATE SHIPMENT DATA FETCHED FROM GENERATE EXCEL DATA: {data}')
  """
  data: {
    order1: {
      "Customer Info": {
        Name, email, phone, notes, etc.
      }
      "Packages": {
        "Box 1": {
          units, length, weight, batteries, etc.
        }
      }
    }
  }
  """
    # format:
    # data["Headings"] = headings
      # headings["Main Headings"]
      # headings["Package Headings"] (return lists)

    # data[order] = order_info
      # order_info["Customer Info"] = customer
      # order_info["Packages"] = pk_details
        # pk_details["Length"] = pk.dimensions[0] etc.

      # region define headings
  customer_headings = ["Customer Info", ""]
  pk_headings = ["Packages", "Units", "Length", "Width", "Height", "Gross Weight (kg)", "CBM", "Has Batteries", "Package Cost"]

    # note: inefficient, change this later
    # False = do not resize
    # "" == resize by looping through each cell
    # str == resize based on len(str)
    # index of each heading is assumed to be its position in the sheet
  def resize_columns():
    len_arrays = ["Wants insurance", "", False, False, False, False, False, "Gross Weight", "Has Batteries", "Package Cost"]
    for i, hd in enumerate(len_arrays):
      col = ws[get_column_letter(i+1)]
      if hd == False:
        continue
      elif hd == "":
        max_len = 0
        for cell in col:
          try:
            if len(str(cell.value)) > max_len:
              max_len = len(cell.value)
          except:
            pass
        new_width = (max_len + 2) * 1.1    # accounts for padding/pixels
        ws.column_dimensions[get_column_letter(i+1)].width = new_width
      else:
        new_width = (len(len_arrays[i]) + 2) * 1.1    # accounts for padding/pixels
        ws.column_dimensions[get_column_letter(i+1)].width = new_width

  def format_top(row_num):
    border_colour = Color(rgb="000000")
    border_style = Side(style='medium', color=border_colour)
    row = ws[row_num]
    for c in row:
      c.border = Border(top=border_style)

  ws.append(customer_headings)

  pk_headings_start = 3
  for hd, i in zip(pk_headings, range(pk_headings_start, pk_headings_start+len(pk_headings))):
    ws.cell(1, i).value = hd
  
  # bold headings
  for col in range(1, len(customer_headings)+len(pk_headings)):
    ws.cell(1, col).font = Font(bold = True, italic = True, color="00000000")
  # endregion

  # updates for every order
  ws_pointer = 2  # start writing from this row (Name)
  for order in data.keys():
    inc = 0   # pointer relative to ws_pointer
      # unpack order info
    order_info = data[order]  # keys are the same as the headings

      # region write customer info
    customer = order_info["Customer Info"]
    cost_row = ws_pointer + 5   # which row is the total cost
    for heading in customer.keys():
      ws.cell(ws_pointer+inc, 1).value = heading
      ws.cell(ws_pointer+inc, 2).value = customer[heading]
      inc += 1
    # endregion

    # region write package info
    inc = 0
    package = order_info["Packages"]
    for pk_id in package.keys():  # pk_id is a string (eg. Box 1)
      ws.cell(ws_pointer+inc, pk_headings_start).value = pk_id  # label
      ws.cell(ws_pointer+inc, pk_headings_start).font = Font(bold = True, color="00000000")

      pk_details = package[pk_id]
      start_col = pk_headings_start+1
      weight_cols = (8, 9)  # weight and CBM are stored here
      this_row = ws_pointer+inc
      this_col = start_col+len(pk_details)
      battery_answers = ("No", "Yes")   # keeps it consistent
      for hd, j in zip(pk_details.keys(), range(0, len(pk_details))):
        val = pk_details[hd]
        
        if val == "cbm":
          dim_col = 5   # dimensions start here
          dims = (f"{get_column_letter(dim_col)}{ws_pointer+inc}", 
                  f"{get_column_letter(dim_col+1)}{ws_pointer+inc}", 
                  f"{get_column_letter(dim_col+2)}{ws_pointer+inc}")
          if pk_details["Units"] == "INCH":
            dims = tuple(v+"*2.54 " for v in dims)
          elif pk_details["Units"] == "CM":
            pass
          else:
            raise ValueError("Invalid measurement unit. Accepted units: INCH, CM")

          val = f"={dims[0]}*{dims[1]}*{dims[2]}/6000"
          cbm_int = pk_details["Length"]*pk_details["Width"]*pk_details["Height"]/6000
          if pk_details["Units"] == "INCH":
            cbm_int = cbm_int * 2.54 * 2.54 * 2.54
          gw_int = pk_details["Gross Weight (kg)"]
        
        if hd == "Has Batteries":
          val = battery_answers[1] if val else battery_answers[0]
        ws.cell(this_row, start_col+j).value = val
        
        # region calculate costs
      # minimum chargeable weight is 10kg
      this_col = start_col+len(pk_details)    # start from cost details
      # gw = ws.cell(this_row, weight_cols[0]).value
      # cbm = ws.cell(this_row, weight_cols[1]).value
      pricewb = 14
      pricewob = 13
      if max(gw_int, cbm_int) < 10:
        if pk_details["Has Batteries"]:
          pcost = 10 * pricewb
        else:
          pcost = 10 * pricewob
      else:
        gw = f"{get_column_letter(weight_cols[0])}{this_row}"
        cbm = f"{get_column_letter(weight_cols[1])}{this_row}"
        ew = f"MAX({gw},{cbm})"    # effective weight
          # =IF(J2="No", Prices!A2*K2, IF(J2="Yes", Prices!B2*K2, "Invalid Input"))
        pcost = f"=IF({get_column_letter(this_col-1)}{this_row}=\"{battery_answers[0]}\", "\
                f"{pricewob}*{ew}, "\
                f"IF({get_column_letter(this_col-1)}{this_row}=\"{battery_answers[1]}\", "\
                f"{pricewb}*{ew}, "\
                f"\"Invalid Input—{battery_answers[1]} or {battery_answers[0]}\"))"
      ws.cell(this_row, this_col).value = pcost
      # endregion
      inc += 1
    
      # get total cost
      ws.cell(cost_row, 2).value = f"=SUM({get_column_letter(this_col)}{ws_pointer}:{get_column_letter(this_col)}{ws_pointer+inc-1})"

      # formatting
    format_top(ws_pointer)
    ws_pointer += max(inc, len(customer))
    df_height = 15    # default row height
    ws.row_dimensions[ws_pointer-1].height = df_height + 5
    # endregion

    #grades = list(data[person].values())
    #ws.append([person] + grades)

  # formatting
  resize_columns()
  ws.freeze_panes = "A1"
  border_colour = Color(rgb="000000")
  border_style = Side(style='medium', color=border_colour)
  col = ws[get_column_letter(pk_headings_start)]
  
  for c in col:
    # need to specify that top remains unchanged
    # otherwise top border will be missing from Packages column
    c.border = Border(top=c.border.top, left=border_style)
  

def write_to_sheet(name, data, file_path=""):
  overwrite = True
  wb = initialize_workbook(name, overwrite)
  populate_shipment(wb, data)

  if file_path == "":
    module_directory = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(module_directory, "temp_files", name)

  save_data(wb, name, file_path, overwrite)    # this determines which file_path to save to
  return file_path